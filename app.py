from flask import Flask, render_template, request, redirect, url_for, session, send_file
import sqlite3
from datetime import date
import openpyxl
import smtplib
from email.mime.text import MIMEText
from apscheduler.schedulers.background import BackgroundScheduler
import os

app = Flask(__name__)
app.secret_key = "chowdary8222"


# ================= DB =================
def get_db():
    conn = sqlite3.connect("attendance.db")
    conn.row_factory = sqlite3.Row
    return conn


# ================= CREATE TABLES =================
def create_tables():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS students (
        student_id TEXT PRIMARY KEY,
        name TEXT,
        email TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS faculty (
        faculty_code TEXT PRIMARY KEY,
        name TEXT,
        password TEXT,
        subject TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS attendance (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        student_id TEXT,
        faculty_code TEXT,
        subject TEXT,
        date TEXT,
        hour TEXT,
        status TEXT
    )
    """)

    cursor.execute("""
    CREATE TABLE IF NOT EXISTS email_logs (
        student_id TEXT,
        last_sent TEXT
    )
    """)

    db.commit()


create_tables()


# ================= EMAIL =================
def send_email(to_email, name, percentage):
    sender = "lakshmanchowdary2005@gmail.com"
    password = "vhmwwqjwsqdrwlhu"

    msg = MIMEText(f"""
Dear {name},

Your attendance is {percentage}%.

This is below 75%.

Please attend classes regularly.

Regards,
HOD of IT Department
""")

    msg['Subject'] = "Low Attendance Alert"
    msg['From'] = sender
    msg['To'] = to_email

    try:
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender, password)
        server.send_message(msg)
        server.quit()
    except Exception as e:
        print("Email Error:", e)


# ================= EMAIL ONCE =================
def send_email_once(student_id, email, name, percentage):
    db = get_db()
    cursor = db.cursor()

    today = str(date.today())

    cursor.execute("SELECT * FROM email_logs WHERE student_id=?", (student_id,))
    record = cursor.fetchone()

    if record and record["last_sent"] == today:
        return

    send_email(email, name, percentage)

    if record:
        cursor.execute("UPDATE email_logs SET last_sent=? WHERE student_id=?",
                       (today, student_id))
    else:
        cursor.execute("INSERT INTO email_logs VALUES (?,?)",
                       (student_id, today))

    db.commit()


# ================= DAILY EMAIL =================
def send_daily_alerts():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM students")
    students = cursor.fetchall()

    for student in students:
        sid = student["student_id"]

        cursor.execute("SELECT COUNT(*) as total FROM attendance WHERE student_id=?", (sid,))
        total = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) as present FROM attendance WHERE student_id=? AND status='Present'", (sid,))
        present = cursor.fetchone()["present"]

        percentage = (present / total * 100) if total else 0

        if percentage < 75 and student["email"]:
            send_email_once(sid, student["email"], student["name"], round(percentage, 2))


# ================= SCHEDULER =================
scheduler = BackgroundScheduler()
scheduler.add_job(send_daily_alerts, 'cron', hour=9, minute=0)
scheduler.start()


# ================= HOME =================
@app.route("/")
def home():
    return render_template("index.html")


# ================= LOGIN =================
@app.route("/login", methods=["GET","POST"])
def login():
    if request.method == "POST":
        code = request.form["code"]
        password = request.form["password"]

        db = get_db()
        cursor = db.cursor()

        if code == "admin" and password == "admin123":
            session["role"] = "admin"
            return redirect("/dashboard")

        cursor.execute("SELECT * FROM faculty WHERE faculty_code=? AND password=?", (code, password))
        faculty = cursor.fetchone()

        if faculty:
            session["role"] = "faculty"
            session["faculty_code"] = code
            return redirect("/dashboard")

        return render_template("login.html", error="Invalid Login")

    return render_template("login.html")


# ================= DASHBOARD =================
@app.route("/dashboard")
def dashboard():
    if "role" not in session:
        return redirect("/login")

    db = get_db()
    cursor = db.cursor()

    faculty = None
    if session["role"] == "faculty":
        cursor.execute("SELECT * FROM faculty WHERE faculty_code=?", (session["faculty_code"],))
        faculty = cursor.fetchone()

    return render_template("dashboard.html", faculty=faculty)


# ================= MARK ATTENDANCE =================
@app.route("/mark_attendance", methods=["GET","POST"])
def mark_attendance():
    if session.get("role") != "faculty":
        return redirect("/login")

    db = get_db()
    cursor = db.cursor()

    cursor.execute("SELECT * FROM faculty WHERE faculty_code=?", (session["faculty_code"],))
    faculty = cursor.fetchone()

    cursor.execute("SELECT student_id,name FROM students ORDER BY student_id")
    students = cursor.fetchall()

    if request.method == "POST":
        hour = request.form["hour"]
        absent = request.form.getlist("absent")

        for s in students:
            sid = s["student_id"]
            status = "Absent" if sid in absent else "Present"

            cursor.execute(
                "SELECT 1 FROM attendance WHERE student_id=? AND subject=? AND date=? AND hour=?",
                (sid, faculty["subject"], str(date.today()), hour)
            )
            exists = cursor.fetchone()

            if not exists:
                cursor.execute("""
                    INSERT INTO attendance
                    (student_id, faculty_code, subject, date, hour, status)
                    VALUES (?,?,?,?,?,?)
                """, (sid, session["faculty_code"], faculty["subject"], str(date.today()), hour, status))

        db.commit()
        return render_template("success.html")

    return render_template("mark_attendance.html", students=students)


# ================= STUDENT REPORT =================
@app.route("/student_report", methods=["GET","POST"])
def student_report():
    db = get_db()
    cursor = db.cursor()

    if request.method == "POST":
        sid = request.form["student_id"].upper()

        cursor.execute("SELECT * FROM students WHERE student_id=?", (sid,))
        student = cursor.fetchone()

        if not student:
            return render_template("student_report.html", error="Student not found", subjects=[], present_counts=[])

        cursor.execute("SELECT COUNT(*) as total FROM attendance WHERE student_id=?", (sid,))
        total = cursor.fetchone()["total"]

        cursor.execute("SELECT COUNT(*) as present FROM attendance WHERE student_id=? AND status='Present'", (sid,))
        present = cursor.fetchone()["present"]

        cursor.execute("""
            SELECT subject,
            SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
            FROM attendance
            WHERE student_id=?
            GROUP BY subject
        """, (sid,))
        data = cursor.fetchall()

        subjects = []
        present_counts = []

        for d in data:
            subjects.append(d["subject"])
            present_counts.append(d["present"] if d["present"] else 0)

        percentage = (present / total * 100) if total else 0

        if percentage < 75 and student["email"]:
            send_email_once(sid, student["email"], student["name"], round(percentage, 2))

        return render_template("student_report.html",
                               student=student,
                               total=total,
                               present=present,
                               percentage=round(percentage, 2),
                               subjects=subjects,
                               present_counts=present_counts)

    return render_template("student_report.html", subjects=[], present_counts=[])


# ================= EXPORT =================
@app.route("/export")
def export():
    db = get_db()
    cursor = db.cursor()

    month = request.args.get("month")

    if not month:
        month = str(date.today().month)

    cursor.execute("""
        SELECT DISTINCT subject 
        FROM attendance 
        WHERE strftime('%m', date)=?
    """, (month.zfill(2),))
    subjects = [row["subject"] for row in cursor.fetchall()]

    cursor.execute("SELECT student_id, name FROM students ORDER BY student_id")
    students = cursor.fetchall()

    wb = openpyxl.Workbook()
    sheet = wb.active

    header = ["Hallticket No"] + subjects + ["Total", "Percentage"]
    sheet.append(header)

    for student in students:
        sid = student["student_id"]
        row = [sid]

        total_all = 0
        present_all = 0

        for subject in subjects:
            cursor.execute("""
                SELECT COUNT(*) as total,
                    SUM(CASE WHEN status='Present' THEN 1 ELSE 0 END) as present
                FROM attendance
                WHERE student_id=? AND subject=? AND strftime('%m', date)=?
            """, (sid, subject, month.zfill(2)))

            res = cursor.fetchone()
            total = res["total"] or 0
            present = res["present"] or 0

            row.append(f"{present}/{total}")
            total_all += total
            present_all += present

        row.append(f"{present_all}/{total_all}")
        percentage = (present_all / total_all * 100) if total_all else 0
        row.append(round(percentage, 2))

        sheet.append(row)

        # ✅ COLOR PART
        from openpyxl.styles import PatternFill

        red = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
        green = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")

        percent_col = len(subjects) + 3
        cell = sheet.cell(row=sheet.max_row, column=percent_col)

        if percentage < 75:
            cell.fill = red
        else:
            cell.fill = green

    file = "attendance.xlsx"
    wb.save(file)

    return send_file(file, as_attachment=True)


# ================= LOGOUT =================
@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# ================= RUN =================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)